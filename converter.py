"""
converter.py
────────────
Docling으로 문서를 파싱하고, 이미지 요소는 LLM으로 분석하여
DocumentElement 리스트로 반환.

지원 포맷: PDF, DOCX, PPTX/PPT, XLSX/XLS, MD, TXT
MPS 가속: Apple Silicon (M1/M2/M3/M4) 자동 감지

포맷별 처리 전략:
  PDF/DOCX/PPTX/XLSX/MD → Docling 직접 처리
  PPT                   → LibreOffice로 PPTX 변환 후 Docling 전달
  XLS                   → xlrd + openpyxl로 XLSX 변환 후 Docling 전달
  TXT                   → 내용을 임시 .md 파일로 저장 후 Docling 전달
"""

import asyncio
import io
import os
import re
import shutil
import tempfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

# MPS fallback: MPS에서 미지원 연산은 CPU로 자동 전환
os.environ.setdefault("PYTORCH_ENABLE_MPS_FALLBACK", "1")

from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import PdfPipelineOptions
from docling.document_converter import DocumentConverter, PdfFormatOption

from llm_router import LLMClient
from models import DocumentElement, ElementType, ParseResponse

_docling_executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="docling-worker")

IMAGE_ANALYSIS_PROMPT = (
    "이 이미지는 문서에서 추출된 것입니다.\n"
    "이미지의 내용을 상세히 설명해주세요.\n"
    "- 차트/그래프: 데이터 수치와 트렌드를 구체적으로 서술\n"
    "- 다이어그램/플로우차트: 구성 요소와 관계를 설명\n"
    "- 표/데이터: 핵심 수치와 의미를 요약\n"
    "- 일반 이미지: 주요 내용과 맥락을 서술\n"
    "설명은 한국어로 작성해주세요."
)


# ─────────────────────────────────────────────────────────────────────────────
# Docling Converter 초기화 (MPS 가속 포함)
# ─────────────────────────────────────────────────────────────────────────────
def _build_docling_converter() -> DocumentConverter:
    import torch

    is_mps = torch.backends.mps.is_available()
    device_label = "MPS (Apple Silicon)" if is_mps else "CPU"
    print(f"[DocParser] PyTorch device: {device_label}")

    pdf_options = PdfPipelineOptions()
    pdf_options.generate_picture_images = True

    try:
        from docling.datamodel.pipeline_options import AcceleratorDevice, AcceleratorOptions

        pdf_options.accelerator_options = AcceleratorOptions(
            device=AcceleratorDevice.MPS if is_mps else AcceleratorDevice.CPU,
        )
        print(f"[DocParser] Docling accelerator: {pdf_options.accelerator_options.device.value}")
    except (ImportError, AttributeError):
        print("[DocParser] Docling AcceleratorOptions 미지원 버전 → PyTorch 자동 감지 모드")

    return DocumentConverter(
        allowed_formats=[
            InputFormat.PDF,
            InputFormat.DOCX,
            InputFormat.PPTX,
            InputFormat.XLSX,
            InputFormat.MD,
        ],
        format_options={
            InputFormat.PDF: PdfFormatOption(pipeline_options=pdf_options),
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 포맷 정규화: PPT/XLS/TXT → Docling이 읽을 수 있는 포맷으로 변환
# Returns: (normalized_path, is_temp)
#   is_temp=True 이면 호출자가 삭제 책임
# ─────────────────────────────────────────────────────────────────────────────
def _normalize_to_docling_format(file_path: str) -> tuple[str, bool]:
    suffix = Path(file_path).suffix.lower()

    # ── TXT → 임시 .md 파일 ──────────────────────────────────────────────────
    if suffix == ".txt":
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".md", mode="w", encoding="utf-8")
        tmp.write(content)
        tmp.close()
        return tmp.name, True

    # ── PPT → PPTX (LibreOffice 필요) ────────────────────────────────────────
    if suffix == ".ppt":
        tmp_dir = tempfile.mkdtemp()
        exit_code = os.system(
            f'libreoffice --headless --convert-to pptx "{file_path}" --outdir "{tmp_dir}"'
        )
        if exit_code != 0:
            shutil.rmtree(tmp_dir, ignore_errors=True)
            raise RuntimeError(
                ".ppt 변환 실패. LibreOffice 설치 필요: brew install --cask libreoffice"
            )
        converted = Path(tmp_dir) / (Path(file_path).stem + ".pptx")
        return str(converted), True

    # ── XLS → XLSX (xlrd + openpyxl) ─────────────────────────────────────────
    if suffix == ".xls":
        try:
            import xlrd
            import openpyxl
        except ImportError:
            raise RuntimeError(".xls 변환에 xlrd가 필요합니다: uv add xlrd")

        wb_old = xlrd.open_workbook(file_path)
        wb_new = openpyxl.Workbook()
        wb_new.remove(wb_new.active)

        for sheet_name in wb_old.sheet_names():
            ws_old = wb_old.sheet_by_name(sheet_name)
            ws_new = wb_new.create_sheet(title=sheet_name)
            for row in range(ws_old.nrows):
                for col in range(ws_old.ncols):
                    ws_new.cell(row=row + 1, column=col + 1, value=ws_old.cell_value(row, col))

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb_new.save(tmp.name)
        tmp.close()
        return tmp.name, True

    # ── 나머지 (PDF/DOCX/PPTX/XLSX/MD) 는 그대로 ────────────────────────────
    return file_path, False


# ─────────────────────────────────────────────────────────────────────────────
# Document Parser Service
# ─────────────────────────────────────────────────────────────────────────────
class DocumentParserService:
    """
    FastAPI lifespan에서 1회 초기화 → 요청마다 재사용.
    Docling 모델 로딩(수백 MB)을 매 요청마다 반복하지 않기 위함.
    """

    def __init__(self):
        self._converter = _build_docling_converter()

    async def parse(
        self,
        file_path: str,
        filename: str,
        llm_client: LLMClient,
    ) -> ParseResponse:
        # 1. 포맷 정규화 (PPT/XLS/TXT 전처리)
        normalized_path, is_temp = _normalize_to_docling_format(file_path)

        try:
            # 2. Docling 변환 (CPU/GPU 바운드 → thread executor)
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                _docling_executor,
                self._converter.convert,
                normalized_path,
            )
        finally:
            # 임시 파일/디렉토리 정리
            if is_temp:
                tmp_dir = os.path.dirname(normalized_path)
                # PPT 변환 시 tmpdir 전체 삭제, 나머지는 파일만 삭제
                if Path(file_path).suffix.lower() == ".ppt":
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                elif os.path.exists(normalized_path):
                    os.unlink(normalized_path)

        doc = result.document
        elements: list[DocumentElement] = []

        # 3. 문서 요소 순회
        for item, _level in doc.iterate_items():
            class_name = type(item).__name__
            page_no = _extract_page_number(item)

            if class_name == "PictureItem":
                element = await self._handle_picture(item, doc, page_no, llm_client)
                if element:
                    elements.append(element)

            elif class_name == "TableItem":
                element = _handle_table(item, page_no)
                if element:
                    elements.append(element)

            elif class_name in ("TextItem", "SectionHeaderItem", "ListItem"):
                element = _handle_text(item, page_no)
                if element:
                    elements.append(element)

        return ParseResponse(
            filename=filename,
            provider_used=f"{llm_client.provider} / {llm_client.model_name}",
            total_elements=len(elements),
            elements=elements,
        )

    async def _handle_picture(
        self,
        item,
        doc,
        page_no: int | None,
        llm_client: LLMClient,
    ) -> DocumentElement | None:
        try:
            image = item.get_image(doc)
            if image is None:
                return None

            buf = io.BytesIO()
            image.save(buf, format="PNG")
            image_bytes = buf.getvalue()

            description = await llm_client.analyze_image(image_bytes, IMAGE_ANALYSIS_PROMPT)

            return DocumentElement(
                page_number=page_no,
                type=ElementType.IMAGE,
                content=f"[이미지 분석 - {llm_client.model_name}]\n\n{description}",
            )
        except Exception as e:
            return DocumentElement(
                page_number=page_no,
                type=ElementType.IMAGE,
                content=f"[이미지 분석 실패: {e}]",
            )



# ─────────────────────────────────────────────────────────────────────────────
# 콘텐츠 필터링
# ─────────────────────────────────────────────────────────────────────────────
_MEANINGLESS_PATTERN = re.compile(r'^[\s\W_]+$')
_MIN_CONTENT_LENGTH = 2


def _is_meaningful(text: str) -> bool:
    cleaned = text.strip()
    if len(cleaned) < _MIN_CONTENT_LENGTH:
        return False
    if _MEANINGLESS_PATTERN.match(cleaned):
        return False
    return True


def _clean_content(text: str) -> str:
    text = re.sub(r' {2,}', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


# ─────────────────────────────────────────────────────────────────────────────
# 순수 함수 헬퍼
# ─────────────────────────────────────────────────────────────────────────────
def _extract_page_number(item) -> int | None:
    try:
        if hasattr(item, "prov") and item.prov:
            return item.prov[0].page_no
    except (IndexError, AttributeError):
        pass
    return None


def _handle_table(item, page_no: int | None) -> DocumentElement | None:
    try:
        md = item.export_to_markdown()
        if _is_meaningful(md):
            return DocumentElement(
                page_number=page_no,
                type=ElementType.TABLE,
                content=_clean_content(md),
            )
    except Exception:
        pass
    return None


def _handle_text(item, page_no: int | None) -> DocumentElement | None:
    text = getattr(item, "text", "").strip()
    if _is_meaningful(text):
        return DocumentElement(
            page_number=page_no,
            type=ElementType.TEXT,
            content=_clean_content(text),
        )
    return None
